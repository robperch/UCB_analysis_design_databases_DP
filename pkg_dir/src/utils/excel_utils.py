## MODULE WITH UTIL FUNCTIONS - MANAGING EXCEL FILES





"----------------------------------------------------------------------------------------------------------------------"
####################################################### Imports ########################################################
"----------------------------------------------------------------------------------------------------------------------"


"--- Standard library imports ---"


"--- Third party imports ---"

import pandas as pd
from openpyxl import load_workbook


## Local application imports





"----------------------------------------------------------------------------------------------------------------------"
####################################################### Functions ######################################################
"----------------------------------------------------------------------------------------------------------------------"


## Creating excel writer to generate spreadsheet
def excel_writer(res_excel_path, res_excel_name):
    """
    Creating excel writer to generate spreadsheet

    :param res_excel_path (string): path to where the resulting excel file will be saved after generated
    :param res_excel_name (string): name of the resulting excel file
    :return writer (?): excel writer object that responsible of generating the spreadsheet
    :return excel_formats (dictionary): dict with the formats that will be used for the resulting file
    """


    ## Creating writer
    writer = pd.ExcelWriter(res_excel_path + res_excel_name, engine="xlsxwriter")

    ## Creating workbook to create formats used in columns
    workbook = writer.book

    ## Creating doctionary of formats for workbook
    excel_formats = {
        "number": workbook.add_format({"num_format": "#,#0"}),
        "currency": workbook.add_format({"num_format": "$#,###.#0"}),
        "percent": workbook.add_format({"num_format": "#.#0%"}),
    }


    return writer, excel_formats



## Writing data on a specific tab of a file
def write_df_in_excel(dfx, writer, excel_formats, worksheet_name):
    """
    Writing final summary dataframe in excel output file

    :param dfx (dataframe): df created for this specific tab and file
    :param writer (?): writer object to print results into excel file and format them
    :param excel_formats (dict): format structures used for the columns
    :param worksheet_name (string): name of the tab where the results will be printed
    :return None:
    """



    ## Writing dataframe into excel document
    dfx.to_excel(writer, sheet_name=worksheet_name)

    ## Worksheet object to allow formatting
    worksheet = writer.sheets[worksheet_name]

    ## Formatting worksheet columns
    worksheet.set_column("A:A", 10, None) ## Index
    worksheet.set_column("B:B", 17, None) ## Ranking
    worksheet.set_column("C:C", 24.5, None) ## Cuenta_contabilizar
    worksheet.set_column("D:D", 17, excel_formats["currency"]) ## Gasto_prom
    worksheet.set_column("E:E", 17, excel_formats["percent"]) ## Prop_total
    worksheet.set_column("F:F", 17, excel_formats["currency"]) ## Total
    worksheet.set_column("G:G", 17, excel_formats["number"]) ## Uuid


    return



## Template function: Writing data on an excel file
def generate_save_excel(dfx):
    """
    Writing data on an excel file

    :param dfx (dataframe): df that will be inserted into de excel file
    :return None:
    """


    ## Defining function to write a dataframe into an excel file
    def excel_from_df(dfx, writer, excel_formats, worksheet_name):
        """
        Defining function to write a dataframe into an excel file

        :param dfx (dataframe): df created for this specific tab and file
        :param writer (?): writer object to print results into excel file and format them
        :param excel_formats (dict): format structures used for the columns
        :param worksheet_name (string): name of the tab where the results will be printed
        :return None:
        """

        ## Writing dataframe into excel document
        dfx.to_excel(writer, sheet_name=worksheet_name)

        ## Worksheet object to allow formatting
        worksheet = writer.sheets[worksheet_name]

        ## Formatting worksheet columns
        worksheet.set_column("A:A", 10, None)  ## Col_1
        worksheet.set_column("B:B", 17, None)  ## Col_2
        worksheet.set_column("C:C", 24.5, None)  ## Col_3
        worksheet.set_column("D:D", 17, excel_formats["currency"])  ## Col_4
        worksheet.set_column("E:E", 17, excel_formats["percent"])  ## Col_5
        worksheet.set_column("F:F", 17, excel_formats["currency"])  ## Col_6
        worksheet.set_column("G:G", 17, excel_formats["number"])  ## Col_7


        return


    ## Generating writer and excel formats
    res_excel_path = "foo"
    res_excel_name = "bar"
    writer, excel_formats = excel_writer(res_excel_path, res_excel_name)

    ## Writing
    excel_from_df(dfx, writer, excel_formats, "Name_of_tab")

    ## Saving excel file
    writer.save()


    return



## Function to read obtain the sheet/tab names in the file
def get_sheet_names_xlsx(file_path):
    """
    Function to read obtain the sheet/tab names in the file

    :param file_path (string): path to excel file that will be loaded
    :return sheet_names (list): list will all the names of the excel file sheets
    """


    ## Loading workbook
    wb = load_workbook(file_path, read_only=True, keep_links=False)

    ## Obtainig the sheet names
    sheet_names = wb.sheetnames


    return sheet_names




"----------------------------------------------------------------------------------------------------------------------"
"----------------------------------------------------------------------------------------------------------------------"
## END OF FILE ##
"----------------------------------------------------------------------------------------------------------------------"
"----------------------------------------------------------------------------------------------------------------------"
